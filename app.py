import sys
import openpyxl
from interface.design import *
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog


class Filter(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)
        self.btnOpenFile.clicked.connect(self.open_plan)
        self.btnSetColumn.clicked.connect(self.get_column)
        self.btnFilter.clicked.connect(self.filtrar)
        self.btnRest.clicked.connect(self.get_others)
        self.btnSaveFile.clicked.connect(self.save_plan)

    def open_plan(self):
        file, _ = QFileDialog.getOpenFileName(
            self.centralwidget,
            'Abrir arquivo',
            )
        self.inputChooseFile.setText(file)


        global wb
        wb = openpyxl.load_workbook(file)

        global plan1
        plan1 = wb[wb.sheetnames[0]]

    def get_column(self):
        global coluna
        coluna = self.inputColuna.text()


    def filtrar(self):
        global plan1
        variantes = list(self.inputVariantes.text().split(','))
        termo = self.inputTermo.text()
        plan2 = wb.create_sheet(termo)
        termos = variantes
        termos.append(termo)

        global col_filter
        col_filter = list(plan1[coluna])
        index = 1
        global col_priority
        col_priority = plan1['AX']


        for i in range(1, len(col_filter) - 1):
            for j in range(len(termos)):
                if (col_filter[i].value.lower()).find(termos[j].lower()) >= 0 and col_priority[i].value != 'x':
                    plan2.cell(index, 1).value = col_filter[i].value
                    index += 1
                    col_priority[i].value = 'x'
                else:
                    pass
        index = 1


    def get_others(self):

        index = 1
        plan3 = wb.create_sheet('Others')
        for i in range(1, len(col_filter) - 1):
            if col_priority[i].value != 'x':
                plan3.cell(index, 1).value = col_filter[i].value
                index += 1


    def save_plan(self):
        plan, _ = QFileDialog.getSaveFileName(
            self.centralwidget,
            'Salvar planilha',
           )
        wb.save(plan + '.xlsx')

if __name__ == '__main__':
    qt = QApplication(sys.argv)
    filter = Filter()
    filter.show()
    qt.exec_()
